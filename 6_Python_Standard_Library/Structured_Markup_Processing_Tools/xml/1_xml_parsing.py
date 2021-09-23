"""
Prerequisites:
https://www.geeksforgeeks.org/xml-parsing-python/

Requirements:
In this program we are going to implement below things
1) Download xml from web
2) Parse that xml file
3) Save it in csv format
"""

# To handle url
import requests

# To works with csv files
import csv

# To work with spreadsheets
# To install: pip install openpyxl
import openpyxl
from openpyxl.styles import Font

# To parse xml files
import xml.etree.ElementTree as ET


def download_xml():
    """
    This function reads the xml data from the web and save it to an xml file
    locally
    TODO: Download xml file on a dynamic path
    """
    url = "https://www.hindustantimes.com/rss/topnews/rssfeed.xml"
    try:
        response = requests.get(url)
        # TODO: check whether that file present or not before creating
        # TODO: If that file present create a new file
        with open("top_news.xml", "wb") as top_news:
            top_news.write(response.content)
    except Exception as error:
        print("Something we wrong while downloading xml locally")
        print(error)


def parse_xml(file_name):
    """
    Function for parsing an xml file
    :param file_name:
    :return:
    """

    # Creating ElementTree object by using parse
    tree_object = ET.parse(file_name)

    # Getting root of the tree object as an element
    root = tree_object.getroot()

    # Creating an empty list to all new item dictionaries
    new_items_dict = []

    # Here we are taking all item tags from an xml
    for item in root.findall("./channel/item"):

        # To to store all child tags of item
        news = {}
        for child in item:

            # TODO: study xml namespace concepts
            if child.tag == "{http://search.yahoo.com/mrss/}content":
                news['media'] = child.attrib['url']
            else:
                news[child.tag] = child.text
        new_items_dict.append(news)
    return new_items_dict


def save_as_csv(list_of_dicts, csv_file_name):
    header = ["title", "description", "link", "guid", "pubDate", "media"]
    with open(csv_file_name, "w") as csv_file:
        # Creating writer object
        writer = csv.DictWriter(csv_file, fieldnames=header)

        # Writing header
        writer.writeheader()

        # Writing rows
        writer.writerows(list_of_dicts)


def save_to_spreadsheet(list_of_dicts, file_name):

    header = ["title", "description", "link", "guid", "pubDate", "media"]
    # Creating spreadsheet
    wb = openpyxl.Workbook()

    # A workbook always created at lease with one worksheet
    # We can get it by using below function
    ws = wb.active

    # Giving title to work sheet
    ws.title = "hindustantimes"

    # loading sheet
    my_sheet = wb[ws.title]

    # Adding hearer in the sheet
    for i in range(len(header)):
        my_sheet.cell(row=1, column=i+1).value = header[i]
        my_sheet.cell(row=1, column=i+1).font = Font(size=11, bold=True, name='Times New Roman', italic=True)
        my_cell = openpyxl.utils.get_column_letter(i+1)
        my_sheet.column_dimensions[my_cell].width = len(str(header[i]))+2





    wb.save("hindustan_times.xlsx")


# download_xml()
news_items_list = parse_xml("top_news.xml")
print(news_items_list)
save_as_csv(news_items_list, "latest_news.csv")
save_to_spreadsheet(news_items_list,"hindustan_times" )
